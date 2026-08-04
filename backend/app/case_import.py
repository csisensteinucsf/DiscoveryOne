from __future__ import annotations

import gzip
import json
import os
import re
import shutil
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from . import models
from .safe_log import debug_suppressed as _debug_suppressed
from .hold_workflows import (
    set_search_holds,
    sync_legacy_custodian_to_default_hold,
    sync_search_hold_statuses,
)
from .ticket_provider_labels import external_ticket_label


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _normalize_person_label(value: str | None) -> str:
    text = (value or "").strip().lower()
    if not text:
        return ""
    text = re.sub(r"[^a-z0-9@]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _custodian_matches_claimant(*, claimant: str | None, name: str | None, email: str | None) -> bool:
    claim = _normalize_person_label(claimant)
    if not claim or claim in {"na", "n/a"}:
        return False
    email_norm = _normalize_person_label(email)
    if "@" in claim and email_norm and email_norm == claim:
        return True
    name_norm = _normalize_person_label(name)
    if not name_norm:
        return False
    if name_norm == claim:
        return True
    if len(claim) >= 4 and (claim in name_norm or name_norm in claim):
        return True
    return False


CELL_MAX_LENGTH = int(os.getenv("CASE_IMPORT_CELL_MAX_LENGTH", "2048"))
XLSX_MAX_ROWS = int(os.getenv("CASE_IMPORT_XLSX_MAX_ROWS", "100000"))
XLSX_MAX_COLS = int(os.getenv("CASE_IMPORT_XLSX_MAX_COLS", "256"))
ZIP_MAX_FILES = int(os.getenv("CASE_IMPORT_ZIP_MAX_FILES", "50"))
ZIP_MAX_ENTRY_BYTES = int(os.getenv("CASE_IMPORT_ZIP_MAX_ENTRY_BYTES", str(50 * 1024 * 1024)))
ZIP_MAX_UNCOMPRESSED_BYTES = int(os.getenv("CASE_IMPORT_ZIP_MAX_UNCOMPRESSED_BYTES", str(200 * 1024 * 1024)))
_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize_text(value: str, preserve_case: bool = False) -> str:
    if not value:
        return ""
    cleaned = value.replace("\ufeff", "")
    cleaned = _CONTROL_CHARS.sub("", cleaned)
    cleaned = cleaned.strip()
    if len(cleaned) > CELL_MAX_LENGTH:
        cleaned = cleaned[:CELL_MAX_LENGTH]
    return cleaned


def _clean_cell(value: Any, preserve_case: bool = False) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    text = str(value)
    return _sanitize_text(text, preserve_case=preserve_case)


def _normalize_header(value: str) -> str:
    if not value:
        return ""
    return " ".join(value.strip().lower().split())


def _bool_from_status(text: str) -> bool:
    if not text:
        return False
    val = text.strip().lower()
    if val in {"", "n/a", "na", "none", "no", "none requested"}:
        return False
    keywords = ("hold", "preserv", "request", "restore", "purview", "legal", "sent", "received", "pending")
    return any(token in val for token in keywords)


def _status_flag(text: str) -> bool:
    if not text:
        return False
    val = text.strip().lower()
    return val not in {"", "n/a", "na", "no"}  # any value means work started


def _consent_status(text: str) -> Optional[str]:
    if not text:
        return None
    val = text.strip().lower()
    if "receiv" in val or "complete" in val or val in {"y", "yes"}:
        return "received"
    if "sent" in val or "request" in val:
        return "sent"
    return None


def _parse_date(value: str) -> Optional[str]:
    if not value:
        return None
    from datetime import datetime as _dt

    value = value.strip().replace(".", "")
    fmts = ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"]
    for fmt in fmts:
        try:
            return _dt.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_date_range(text: str) -> Tuple[Optional[str], Optional[str]]:
    if not text:
        return (None, None)
    cleaned = text.replace("–", "-").replace("—", "-").replace(" to ", "-")
    parts = re.split(r"\s*-\s*", cleaned)
    if len(parts) >= 2:
        start = _parse_date(parts[0])
        end = _parse_date(parts[1])
        return (start, end)
    return (None, None)


def _safe_filename(name: str) -> str:
    return os.path.basename(name or "upload.xlsx")


def _case_name_from_filename(filename: str) -> str:
    stem = Path(filename).stem
    return stem.strip() or "Imported Case"


def _color_from_case_name(case_name: str) -> Optional[str]:
    if "-" in case_name:
        return case_name.split("-", 1)[1] or None
    return None


def _normalize_email(email: Optional[str]) -> Optional[str]:
    if not email:
        return None
    email = email.strip().lower()
    return email or None

IMPORTSTATUS_KEEP = int(os.getenv("CASE_IMPORT_LOG_KEEP", "30"))


SEARCH_FIELDS = {"name", "criteria", "date range", "completed", "exported", "delivered"}


def _detect_search_field(header: str) -> Optional[Tuple[str, str]]:
    if not header:
        return None
    parts = header.split()
    slot = None
    field = None
    if parts[0] == "search":
        if len(parts) >= 3 and parts[1].isdigit():
            slot = parts[1]
            field = " ".join(parts[2:])
        else:
            field = " ".join(parts[1:]) if len(parts) > 1 else "name"
    elif parts[0].isdigit() and len(parts) > 1:
        slot = parts[0]
        field = " ".join(parts[1:])
    elif parts[0] in {"exported", "delivered"}:
        slot = "1"
        field = parts[0]
    else:
        return None
    field = field.strip()
    if field not in SEARCH_FIELDS:
        return None
    if slot is None:
        slot = "1"
    return slot, field


RECOGNIZED_COLUMNS = {
    "servicenow case",
    "name",
    "custodian",
    "custodian name",
    "custodian first name",
    "custodian last name",
    "first",
    "last",
    "email",
    "email address",
    "custodian email",
    "email preservation status",
    "box",
    "box preservation status",
    "box preservation",
    "box hold",
    "box preservation status",
    "box hold status",
    "box preservation status",
    "box preservation status",
    "box preservation status",
    "p drive",
    "network drives preservation status",
    "microsoft 365 status",
    "google vault status",
    "slack status",
    "consent",
    "phone",
    "computer",
    "search name",
    "search date range",
    "search criteria",
    "search completed",
    "exported",
    "delivered",
    "search 1 name",
    "search 2 name",
    "search 1 criteria",
    "search 2 criteria",
    "search 1 date range",
    "search 2 date range",
    "legal hold unique id",
    "project name",
    "legal hold name",
    "legal hold status",
    "custodian status",
    "notice version sent",
    "preserved services",
    "failed services",
}


@dataclass
class FileImportResult:
    filename: str
    case_name: Optional[str] = None
    case_id: Optional[int] = None
    status: str = "pending"
    created_case: bool = False
    total_rows: int = 0
    custodians_created: int = 0
    custodians_updated: int = 0
    custodians_skipped: int = 0
    searches_created: int = 0
    searches_updated: int = 0
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    unmapped_fields: set[str] = field(default_factory=set)
    servicenow_case: Optional[str] = None

    def finalize(self) -> None:
        if self.errors:
            self.status = "error"
        elif self.warnings or self.unmapped_fields:
            self.status = "partial"
        else:
            self.status = "success"

    def as_dict(self) -> Dict[str, Any]:
        return {
            "filename": self.filename,
            "case_name": self.case_name,
            "case_id": self.case_id,
            "status": self.status,
            "created_case": self.created_case,
            "total_rows": self.total_rows,
            "custodians_created": self.custodians_created,
            "custodians_updated": self.custodians_updated,
            "custodians_skipped": self.custodians_skipped,
            "searches_created": self.searches_created,
            "searches_updated": self.searches_updated,
            "warnings": self.warnings,
            "errors": self.errors,
            "unmapped_fields": sorted(self.unmapped_fields),
            "servicenow_case": self.servicenow_case,
        }


class CaseSpreadsheetImporter:
    EMAIL_KEYS = {"email address", "email", "custodian email"}
    NAME_KEYS = {"name", "custodian", "custodian name"}
    FIRST_KEYS = {"custodian first name", "first", "first name"}
    LAST_KEYS = {"custodian last name", "last", "last name"}

    def __init__(self, db: Session):
        self.db = db
        root = os.getenv("CASE_IMPORT_REPORT_DIR", "/app/import_reports")
        self.report_root = Path(root)
        self.report_root.mkdir(parents=True, exist_ok=True)
        log_root = os.getenv("LOG_DIR", "/app/logs")
        self.log_dir = Path(log_root)
        self.log_dir.mkdir(parents=True, exist_ok=True)

    def import_uploads(self, uploads: List[Tuple[str, bytes]]) -> Dict[str, Any]:
        run_id = _now().strftime("%Y%m%d-%H%M%S")
        run_dir = self.report_root / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        results: List[Dict[str, Any]] = []
        for original_name, payload in uploads:
            safe_name = _safe_filename(original_name)
            lower = safe_name.lower()
            if lower.endswith(".zip"):
                results.extend(self._process_zip(safe_name, payload))
            elif lower.endswith(".xlsx"):
                results.append(self._process_workbook(safe_name, payload).as_dict())
            else:
                result = FileImportResult(filename=safe_name, status="error")
                result.errors.append("Unsupported file type (expected .xlsx or .zip).")
                result.finalize()
                results.append(result.as_dict())
        report_meta = self._write_report_files(run_dir, results)
        created_cases: List[Dict[str, Any]] = []
        seen_cases: set[int] = set()
        for item in results:
            cid = item.get("case_id")
            if item.get("created_case") and cid and cid not in seen_cases:
                created_cases.append({"id": cid, "name": item.get("case_name")})
                seen_cases.add(cid)
        response = {
            "import_id": run_id,
            "report_dir": str(run_dir),
            "files": results,
            "created_cases": created_cases,
        }
        response.update(report_meta)
        return response

    def _process_zip(self, archive_name: str, payload: bytes) -> List[Dict[str, Any]]:
        outcomes: List[Dict[str, Any]] = []
        try:
            with zipfile.ZipFile(BytesIO(payload)) as zf:
                xlsx_entries = [
                    entry
                    for entry in zf.infolist()
                    if (not entry.is_dir()) and entry.filename.lower().endswith(".xlsx")
                ]
                if ZIP_MAX_FILES > 0 and len(xlsx_entries) > ZIP_MAX_FILES:
                    too_many = FileImportResult(filename=archive_name, status="error")
                    too_many.errors.append(f"Archive contains too many .xlsx files (limit {ZIP_MAX_FILES}).")
                    too_many.finalize()
                    outcomes.append(too_many.as_dict())
                    return outcomes

                total_uncompressed = 0
                for entry in xlsx_entries:
                    file_size = int(getattr(entry, "file_size", 0) or 0)
                    total_uncompressed += file_size
                    if ZIP_MAX_UNCOMPRESSED_BYTES > 0 and total_uncompressed > ZIP_MAX_UNCOMPRESSED_BYTES:
                        limit_mb = max(1, ZIP_MAX_UNCOMPRESSED_BYTES // (1024 * 1024))
                        problem = FileImportResult(filename=archive_name, status="error")
                        problem.errors.append(f"Archive expanded size exceeds limit ({limit_mb} MB).")
                        problem.finalize()
                        outcomes.append(problem.as_dict())
                        break

                    if ZIP_MAX_ENTRY_BYTES > 0 and file_size > ZIP_MAX_ENTRY_BYTES:
                        limit_mb = max(1, ZIP_MAX_ENTRY_BYTES // (1024 * 1024))
                        problem = FileImportResult(filename=f"{archive_name}:{entry.filename}", status="error")
                        problem.errors.append(f"Workbook size exceeds limit ({limit_mb} MB).")
                        problem.finalize()
                        outcomes.append(problem.as_dict())
                        continue

                    with zf.open(entry) as fh:
                        data = fh.read()
                    result = self._process_workbook(f"{archive_name}:{entry.filename}", data)
                    outcomes.append(result.as_dict())
        except Exception as exc:
            problem = FileImportResult(filename=archive_name, status="error")
            problem.errors.append(f"Zip archive could not be read: {exc}")
            problem.finalize()
            outcomes.append(problem.as_dict())
        if not outcomes:
            empty = FileImportResult(filename=archive_name, status="error")
            empty.errors.append("Archive did not contain any .xlsx files.")
            empty.finalize()
            outcomes.append(empty.as_dict())
        return outcomes

    def _process_workbook(self, filename: str, payload: bytes) -> FileImportResult:
        result = FileImportResult(filename=filename)
        case_name = _case_name_from_filename(filename)
        result.case_name = case_name
        row_limit = XLSX_MAX_ROWS
        col_limit = XLSX_MAX_COLS
        wb = None
        try:
            wb = load_workbook(BytesIO(payload), data_only=True, read_only=True)
            ws = wb.active
            iterator = iter(ws.iter_rows(values_only=True))
            header_row = next(iterator, None)
        except Exception as exc:
            self.db.rollback()
            result.errors.append(f"Unable to read workbook: {exc}")
            result.finalize()
            return result

        if not header_row:
            result.errors.append("Workbook is empty.")
            result.finalize()
            try:
                if wb is not None:
                    wb.close()
            except Exception as exc:
                _debug_suppressed("suppressed exception in case_import.py:421", exc)
            return result

        if col_limit > 0:
            header_row = tuple(header_row[:col_limit])

        headers_raw = [(_clean_cell(cell, preserve_case=True) if cell is not None else "") for cell in header_row]
        header_meta = [{"raw": raw, "norm": _normalize_header(raw)} for raw in headers_raw]
        normalized_headers = {meta["norm"] for meta in header_meta if meta["norm"]}

        if not normalized_headers.intersection(self.EMAIL_KEYS):
            result.errors.append("No email column detected; cannot import.")
            result.finalize()
            try:
                if wb is not None:
                    wb.close()
            except Exception as exc:
                _debug_suppressed("suppressed exception in case_import.py:438", exc)
            return result

        structured_rows = []
        data_seen = 0
        for idx, row in enumerate(iterator, start=2):
            data_seen += 1
            if row_limit > 0 and data_seen > row_limit:
                result.warnings.append(f"Stopped after {row_limit} rows (row limit reached).")
                break
            if col_limit > 0:
                row = tuple((row or [])[:col_limit])
            values = {}
            non_blank = False
            for col_idx, cell in enumerate(row):
                value = _clean_cell(cell)
                if value:
                    non_blank = True
                if col_idx < len(header_meta):
                    key = header_meta[col_idx]["norm"]
                    if key:
                        values[key] = value
                        if key not in RECOGNIZED_COLUMNS and value:
                            result.unmapped_fields.add(header_meta[col_idx]["raw"] or key)
                    elif value:
                        result.unmapped_fields.add(header_meta[col_idx]["raw"] or f"Column {col_idx+1}")
            if non_blank:
                structured_rows.append((idx, values))

        if not structured_rows:
            result.errors.append("No data rows found.")
            result.finalize()
            try:
                if wb is not None:
                    wb.close()
            except Exception as exc:
                _debug_suppressed("suppressed exception in case_import.py:474", exc)
            return result

        try:
            if wb is not None:
                wb.close()
        except Exception as exc:
            _debug_suppressed("suppressed exception in case_import.py:481", exc)

        case = (
            self.db.query(models.Case).filter(models.Case.name == case_name).first()
        )
        if case is None:
            case = models.Case(
                name=case_name,
                color=_color_from_case_name(case_name),
            )
            self.db.add(case)
            self.db.flush()
            result.created_case = True
        result.case_id = getattr(case, "id", None)

        existing_custodians = {
            _normalize_email(c.email): c
            for c in self.db.query(models.Custodian).filter(models.Custodian.case_id == case.id).all()
            if c.email
        }
        searches = {
            (s.name or "").strip().lower(): s
            for s in self.db.query(models.Search).filter(models.Search.case_id == case.id).all()
        }

        for row_number, mapping in structured_rows:
            result.total_rows += 1
            email = _normalize_email(
                mapping.get("email address")
                or mapping.get("email")
                or mapping.get("custodian email")
            )
            name = (
                mapping.get("name")
                or mapping.get("custodian")
                or mapping.get("custodian name")
            )
            if not name:
                first = mapping.get("custodian first name") or mapping.get("first") or ""
                last = mapping.get("custodian last name") or mapping.get("last") or ""
                name = " ".join(part for part in (first.strip(), last.strip()) if part)
            if not name and email:
                name = email.split("@")[0]

            if not email and not name:
                result.custodians_skipped += 1
                result.warnings.append(f"Row {row_number}: missing custodian name and email.")
                continue

            if not email:
                result.warnings.append(f"Row {row_number}: missing email; custodian skipped.")
                result.custodians_skipped += 1
                continue

            custodian = existing_custodians.get(email)
            action = "updated"
            if custodian is None:
                custodian = models.Custodian(
                    case_id=case.id,
                    name=name or email,
                    email=email,
                )
                self.db.add(custodian)
                self.db.flush()
                existing_custodians[email] = custodian
                result.custodians_created += 1
                action = "created"
            else:
                result.custodians_updated += 1
                if name and name != custodian.name:
                    custodian.name = name

            email_status = mapping.get("email preservation status") or mapping.get("microsoft 365 status")
            if _bool_from_status(email_status):
                custodian.holds_email = True

            box_status = (
                mapping.get("box")
                or mapping.get("box preservation status")
                or mapping.get("box preservation")
                or mapping.get("box hold")
            )
            if _bool_from_status(box_status):
                custodian.holds_box = True

            drive_status = mapping.get("p drive") or mapping.get("network drives preservation status")
            if _bool_from_status(drive_status):
                custodian.holds_onedrive = True

            slack_status = mapping.get("slack status")
            if _bool_from_status(slack_status):
                custodian.holds_slack = True

            consent_value = mapping.get("consent")
            new_consent = _consent_status(consent_value)
            if new_consent:
                if custodian.consent_status != "received" or new_consent == "received":
                    custodian.consent_status = new_consent

            if _custodian_matches_claimant(
                claimant=getattr(case, "claimant", None),
                name=getattr(custodian, "name", None),
                email=getattr(custodian, "email", None),
            ):
                if (getattr(custodian, "ntp_status", "") or "").strip().lower() != "acknowledged":
                    custodian.ntp_status = "na"
                if (getattr(custodian, "consent_status", "") or "").strip().lower() != "received":
                    custodian.consent_status = "na"

            if _status_flag(mapping.get("search completed") or mapping.get("completed")):
                custodian.search_done = True
            if _status_flag(mapping.get("exported")):
                custodian.export_done = True
            if _status_flag(mapping.get("delivered")):
                custodian.delivered_done = True

            if not result.servicenow_case:
                sn_case = mapping.get("servicenow case")
                if sn_case:
                    result.servicenow_case = sn_case.strip()

            search_slots = self._extract_search_slots(mapping)
            for slot_key, search_data in search_slots.items():
                search_name = search_data.get("name") or f"{case_name}-Search {slot_key}"
                lookup = search_name.strip().lower()
                search = searches.get(lookup)
                if search is None:
                    search = models.Search(
                        case_id=case.id,
                        name=search_name,
                        keywords=search_data.get("criteria"),
                        additional=search_data.get("criteria"),
                        status_search="performed" if search_data.get("completed") else "not performed",
                        status_export="performed" if search_data.get("exported") else "not performed",
                        status_delivery="performed" if search_data.get("delivered") else "not performed",
                    )
                    start, end = _parse_date_range(search_data.get("date_range") or "")
                    search.date_from = start
                    search.date_to = end
                    search.custodian_ids = json.dumps([custodian.id])
                    self.db.add(search)
                    self.db.flush()
                    set_search_holds(self.db, search=search, hold_ids=None)
                    searches[lookup] = search
                    result.searches_created += 1
                else:
                    updated = False
                    if search_data.get("criteria"):
                        search.keywords = search_data.get("criteria")
                        updated = True
                    if search_data.get("completed"):
                        search.status_search = "performed"
                        updated = True
                    if search_data.get("exported"):
                        search.status_export = "performed"
                        updated = True
                    if search_data.get("delivered"):
                        search.status_delivery = "performed"
                        updated = True
                    start, end = _parse_date_range(search_data.get("date_range") or "")
                    if start and not search.date_from:
                        search.date_from = start
                        updated = True
                    if end and not search.date_to:
                        search.date_to = end
                        updated = True
                    ids = []
                    try:
                        ids = json.loads(search.custodian_ids or "[]")
                    except Exception:
                        ids = []
                    if custodian.id not in ids:
                        ids.append(custodian.id)
                        search.custodian_ids = json.dumps(ids)
                        updated = True
                    if updated:
                        sync_search_hold_statuses(self.db, search)
                        result.searches_updated += 1

        # Spreadsheet imports use legacy columns; treat them as updates to the default named hold.
        for imported_custodian in existing_custodians.values():
            sync_legacy_custodian_to_default_hold(self.db, imported_custodian)

        if result.servicenow_case and not case.description:
            case.description = f"{external_ticket_label()} {result.servicenow_case}"

        try:
            self.db.commit()
        except Exception as exc:
            self.db.rollback()
            result.errors.append(f"Database error: {exc}")

        result.finalize()
        return result

    def _extract_search_slots(self, mapping: Dict[str, str]) -> Dict[str, Dict[str, str]]:
        slots: Dict[str, Dict[str, str]] = defaultdict(dict)
        for key, value in mapping.items():
            if not value:
                continue
            match = _detect_search_field(key)
            if not match:
                continue
            slot, field = match
            slots[slot][field if field != "date range" else "date_range"] = value
        return slots

    def _write_report_files(self, directory: Path, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        warnings: List[str] = []
        payload = {
            "generated_at": _now().isoformat(),
            "files": results,
        }
        summary_json = directory / "summary.json"
        try:
            summary_json.write_text(json.dumps(payload, indent=2))
        except OSError as exc:
            warnings.append(f"Unable to write {summary_json.name}: {exc}")
        lines = [f"Import report generated {payload['generated_at']} (UTC)", ""]
        for item in results:
            lines.append(f"- {item.get('filename')} -> case {item.get('case_name') or 'unknown'} [{item.get('status')}]")
            lines.append(f"  Custodians: +{item.get('custodians_created',0)} / updated {item.get('custodians_updated',0)} / skipped {item.get('custodians_skipped',0)}")
            lines.append(f"  Searches: +{item.get('searches_created',0)} / updated {item.get('searches_updated',0)}")
            if item.get("warnings"):
                lines.append("  Warnings:")
                for warn in item["warnings"]:
                    lines.append(f"    • {warn}")
            if item.get("errors"):
                lines.append("  Errors:")
                for err in item["errors"]:
                    lines.append(f"    • {err}")
            if item.get("unmapped_fields"):
                lines.append("  Unmapped fields: " + ", ".join(item["unmapped_fields"]))
            lines.append("")
        summary_text = directory / "summary.txt"
        text_body = "\n".join(lines)
        try:
            summary_text.write_text(text_body)
        except OSError as exc:
            warnings.append(f"Unable to write {summary_text.name}: {exc}")
        log_path_obj = self.log_dir / f"importstatus_{directory.name}.log"
        log_path_str: Optional[str] = None
        try:
            log_path_obj.write_text(text_body)
            log_path_str = str(log_path_obj)
            # Retain only the most recent N importstatus logs (uncompressed)
            try:
                candidates = sorted(
                    self.log_dir.glob("importstatus_*.log"),
                    key=lambda p: p.stat().st_mtime,
                    reverse=True,
                )
                for stale in candidates[IMPORTSTATUS_KEEP:]:
                    try:
                        stale.unlink()
                    except OSError:
                        pass
            except Exception as exc:
                _debug_suppressed("suppressed exception in case_import.py:733", exc)
        except OSError as exc:
            warnings.append(f"Unable to write log file: {exc}")
        for message in warnings:
            print(f"[case import] {message}")
        return {"log_path": log_path_str, "log_text": text_body, "report_warnings": warnings}
