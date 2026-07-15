from __future__ import annotations

import csv
import io
import uuid
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook

from . import case_requests as case_request_core
from . import models
from .auth import current_user as get_current_user
from .permissions import ensure_case_request_access, is_tester

router = APIRouter(prefix="/api/case_requests", tags=["case_requests"])

_CUSTODIAN_NAME_KEYS = {
    "name",
    "full name",
    "custodian",
    "custodian name",
    "custodian full name",
}
_CUSTODIAN_EMAIL_KEYS = {
    "email",
    "email address",
    "emailaddress",
    "custodian email",
}
_CUSTODIAN_NOTES_KEYS = {
    "notes",
    "note",
    "comments",
    "comment",
    "details",
}


def _safe_filename(name: Optional[str]) -> str:
    stem = Path(name or "upload").name
    return stem or "upload.bin"


def _decode_text_table_payload(payload: bytes) -> str:
    """
    Decode a CSV/TSV-like payload into text.
    Accepts UTF-8 (with optional BOM) and UTF-16 (common Excel exports).
    """
    if not payload:
        return ""
    try:
        return payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        pass
    if b"\x00" in payload:
        for enc in ("utf-16", "utf-16-le", "utf-16-be"):
            try:
                return payload.decode(enc)
            except UnicodeDecodeError:
                continue
    try:
        return payload.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail="Uploaded file must be UTF-8 or UTF-16 encoded text") from exc


def _normalize_header(value: Optional[str]) -> str:
    text = (value or "").strip().lower()
    for ch in ("_", "-", ".", ":", "/", "\\"):
        text = text.replace(ch, " ")
    return " ".join(text.split())


def _parse_uploaded_custodians_from_bytes(payload: bytes, filename: Optional[str]) -> List[Dict[str, Any]]:
    if not payload:
        return []

    max_rows = case_request_core.CASE_REQUEST_CUSTODIAN_MAX_ROWS
    max_cols = case_request_core.CASE_REQUEST_CUSTODIAN_MAX_COLS
    ext = (Path(filename or "").suffix or "").lower()
    mime = case_request_core.sniff_mime(payload)

    def _append(name: str, email: str, notes: str) -> None:
        name_val = (name or "").strip()
        if not name_val:
            return
        results.append(
            {
                "name": name_val,
                "email": (email or "").strip(),
                "notes": (notes or "").strip(),
            }
        )

    def _find_idx(headers: list[str], keys: set[str]) -> Optional[int]:
        for idx, header in enumerate(headers):
            if header in keys:
                return idx
        return None

    results: List[Dict[str, Any]] = []

    if ext == ".xlsx" or mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        wb = None
        try:
            wb = load_workbook(filename=BytesIO(payload), data_only=True, read_only=True)
            ws = wb.active
            iterator = iter(ws.iter_rows(values_only=True))
            first_row = next(iterator, None)
        except Exception:
            return []

        if not first_row:
            try:
                if wb is not None:
                    wb.close()
            except Exception as exc:
                case_request_core._debug_suppressed("suppressed exception in case_request_custodian_uploads.py:96", exc)
            return []

        if max_cols > 0:
            first_row = tuple((first_row or [])[:max_cols])

        header_row = [str(c) if c is not None else "" for c in (first_row or [])]
        headers = [_normalize_header(v) for v in header_row]
        header_has_keys = any(h in (_CUSTODIAN_NAME_KEYS | _CUSTODIAN_EMAIL_KEYS | _CUSTODIAN_NOTES_KEYS) for h in headers)

        if header_has_keys:
            name_idx = _find_idx(headers, _CUSTODIAN_NAME_KEYS)
            email_idx = _find_idx(headers, _CUSTODIAN_EMAIL_KEYS)
            notes_idx = _find_idx(headers, _CUSTODIAN_NOTES_KEYS)
            seen = 0
            try:
                for row in iterator:
                    seen += 1
                    if max_rows > 0 and seen > max_rows:
                        raise HTTPException(status_code=413, detail=f"Custodian upload exceeds row limit ({max_rows}).")
                    if not row:
                        continue
                    if max_cols > 0:
                        row = tuple((row or [])[:max_cols])
                    name = str(row[name_idx] if name_idx is not None and name_idx < len(row) else "" or "").strip()
                    email = str(row[email_idx] if email_idx is not None and email_idx < len(row) else "" or "").strip()
                    notes = str(row[notes_idx] if notes_idx is not None and notes_idx < len(row) else "" or "").strip()
                    _append(name, email, notes)
                return results
            finally:
                try:
                    if wb is not None:
                        wb.close()
                except Exception as exc:
                    case_request_core._debug_suppressed("suppressed exception in case_request_custodian_uploads.py:130", exc)

        seen = 0
        try:
            row = first_row
            seen += 1
            if max_rows > 0 and seen > max_rows:
                raise HTTPException(status_code=413, detail=f"Custodian upload exceeds row limit ({max_rows}).")
            if row:
                name = str(row[0] if len(row) > 0 else "" or "").strip()
                email = str(row[1] if len(row) > 1 else "" or "").strip()
                notes = str(row[2] if len(row) > 2 else "" or "").strip()
                _append(name, email, notes)

            for row in iterator:
                seen += 1
                if max_rows > 0 and seen > max_rows:
                    raise HTTPException(status_code=413, detail=f"Custodian upload exceeds row limit ({max_rows}).")
                if not row:
                    continue
                if max_cols > 0:
                    row = tuple((row or [])[:max_cols])
                name = str(row[0] if len(row) > 0 else "" or "").strip()
                email = str(row[1] if len(row) > 1 else "" or "").strip()
                notes = str(row[2] if len(row) > 2 else "" or "").strip()
                _append(name, email, notes)
            return results
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception as exc:
                case_request_core._debug_suppressed("suppressed exception in case_request_custodian_uploads.py:161", exc)

    text = _decode_text_table_payload(payload)
    if not text.strip():
        return []

    sample = text[:8192]
    delimiter = ","
    if sample.count("\t") > sample.count(",") and sample.count("\t") >= 1:
        delimiter = "\t"
    elif sample.count(";") > sample.count(",") and sample.count(";") >= 1:
        delimiter = ";"

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    first_row = next(reader, None)
    if first_row is None:
        return []
    if max_cols > 0:
        first_row = list(first_row[:max_cols])

    header = [_normalize_header(str(c) if c is not None else "") for c in (first_row or [])]
    header_has_keys = any(h in (_CUSTODIAN_NAME_KEYS | _CUSTODIAN_EMAIL_KEYS | _CUSTODIAN_NOTES_KEYS) for h in header)

    if header_has_keys:
        dict_reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

        def _first_val(row: dict, keys: set[str]) -> str:
            for raw_key, raw_val in (row or {}).items():
                if _normalize_header(raw_key) in keys:
                    return (raw_val or "").strip()
            return ""

        seen = 0
        for row in dict_reader:
            seen += 1
            if max_rows > 0 and seen > max_rows:
                raise HTTPException(status_code=413, detail=f"Custodian upload exceeds row limit ({max_rows}).")
            if not row:
                continue
            _append(_first_val(row, _CUSTODIAN_NAME_KEYS), _first_val(row, _CUSTODIAN_EMAIL_KEYS), _first_val(row, _CUSTODIAN_NOTES_KEYS))
        return results

    seen = 0
    row = first_row
    seen += 1
    if max_rows > 0 and seen > max_rows:
        raise HTTPException(status_code=413, detail=f"Custodian upload exceeds row limit ({max_rows}).")
    if row:
        name = str((row[0] if len(row) > 0 else "") or "").strip()
        email = str((row[1] if len(row) > 1 else "") or "").strip()
        notes = str((row[2] if len(row) > 2 else "") or "").strip()
        _append(name, email, notes)

    for row in reader:
        seen += 1
        if max_rows > 0 and seen > max_rows:
            raise HTTPException(status_code=413, detail=f"Custodian upload exceeds row limit ({max_rows}).")
        if not row:
            continue
        name = str((row[0] if len(row) > 0 else "") or "").strip()
        email = str((row[1] if len(row) > 1 else "") or "").strip()
        notes = str((row[2] if len(row) > 2 else "") or "").strip()
        _append(name, email, notes)
    return results


def _validate_custodian_upload_bytes(payload: bytes, filename: Optional[str], *, max_bytes: int) -> None:
    if not payload:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if max_bytes and len(payload) > max_bytes:
        max_mb = max(1, max_bytes // (1024 * 1024))
        raise HTTPException(status_code=413, detail=f"Upload exceeds size limit ({max_mb} MB).")

    ext = (Path(filename or "").suffix or "").lower()
    mime = case_request_core.sniff_mime(payload)
    if ext == ".xlsx" or mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        parsed = _parse_uploaded_custodians_from_bytes(payload, filename)
        if not parsed:
            raise HTTPException(status_code=400, detail="Uploaded .xlsx did not contain any valid custodian rows")
        return

    if mime not in {"text/plain", "text/csv"}:
        raise HTTPException(status_code=415, detail="Unsupported attachment format; expected CSV/TSV/text or .xlsx")

    parsed = _parse_uploaded_custodians_from_bytes(payload, filename)
    if not parsed:
        raise HTTPException(status_code=400, detail="Uploaded file did not contain any valid custodian rows")


def _save_upload(file: UploadFile, *, actor: models.User, request: Request) -> tuple[str, str, int]:
    data = file.file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    try:
        file.file.close()
    except Exception as exc:
        case_request_core._debug_suppressed("suppressed exception in case_request_custodian_uploads.py:263", exc)
    try:
        _validate_custodian_upload_bytes(data, file.filename or "upload", max_bytes=case_request_core.MAX_UPLOAD_BYTES)
    except HTTPException as exc:
        if exc.status_code == 413:
            max_mb = max(1, case_request_core.MAX_UPLOAD_BYTES // (1024 * 1024))
            raise HTTPException(status_code=413, detail=f"Upload exceeds size limit ({max_mb} MB).") from exc
        raise
    case_request_core.scan_payload(data, file.filename or "upload", request=request, actor=actor)
    token = f"{uuid.uuid4().hex}_{_safe_filename(file.filename)}"
    dest = case_request_core.CASE_REQUEST_UPLOAD_DIR / token
    dest.write_bytes(data)
    return (file.filename or "upload", str(dest), len(data))


@router.post("/parse_custodian_file")
async def parse_custodian_file(
    request: Request,
    custodian_file: UploadFile = File(...),
    actor: models.User = Depends(get_current_user),
):
    """
    Parse an uploaded custodian CSV/TSV/text/.xlsx and return normalized rows.
    Intended for the requestor UI to preview custodians and run person lookup.
    """
    if not actor:
        raise HTTPException(status_code=401, detail="Not authenticated")
    ensure_case_request_access(actor)
    if is_tester(actor):
        raise HTTPException(status_code=403, detail="Tester accounts cannot submit requests")
    try:
        payload = await custodian_file.read()
    finally:
        try:
            await custodian_file.close()
        except Exception as exc:
            case_request_core._debug_suppressed("suppressed exception in case_request_custodian_uploads.py:296", exc)
    filename = custodian_file.filename or "custodians"
    if not payload:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if case_request_core.MAX_UPLOAD_BYTES > 0 and len(payload) > case_request_core.MAX_UPLOAD_BYTES:
        max_mb = max(1, case_request_core.MAX_UPLOAD_BYTES // (1024 * 1024))
        raise HTTPException(status_code=413, detail=f"Upload exceeds size limit ({max_mb} MB).")

    case_request_core.scan_payload(payload, filename, request=request, actor=actor)
    custodians = _parse_uploaded_custodians_from_bytes(payload, filename)
    if not custodians:
        raise HTTPException(status_code=400, detail="Uploaded file did not contain any valid custodian rows")
    return {"count": len(custodians), "custodians": custodians}
